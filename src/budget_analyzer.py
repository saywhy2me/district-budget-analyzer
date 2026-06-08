"""
District Budget & Account Variance Analyzer
===========================================

Loads a chart-of-accounts / budget CSV into a SQLite ledger, reconciles and
validates the data, computes budget-vs-actual variance, flags compliance
exceptions, and writes a formatted Excel + CSV report.

Designed to mirror school-district Fiscal Services workflows.

Usage:
    python -m src.budget_analyzer --input data/sample_budget.csv --out reports
    python -m src.budget_analyzer --input data/sample_budget.csv --threshold 0.95
"""

from __future__ import annotations

import argparse
import sqlite3
from dataclasses import dataclass
from pathlib import Path

import pandas as pd

REQUIRED_COLUMNS = ["fund", "account", "description", "budget", "actual"]

# Spending at or above this fraction of budget (but not over) is flagged WATCH.
DEFAULT_WATCH_THRESHOLD = 0.90


# --------------------------------------------------------------------------- #
# Validation
# --------------------------------------------------------------------------- #
class ValidationError(ValueError):
    """Raised when the input ledger fails reconciliation checks."""


def validate(df: pd.DataFrame) -> pd.DataFrame:
    """Validate structure and values; return a cleaned, typed DataFrame.

    Checks performed (the "reconciliation" step):
      * all required columns are present
      * budget and actual parse as numbers
      * no negative budgets or negative actuals
      * no duplicate fund/account lines (which would double-count a balance)
    """
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValidationError(f"Missing required column(s): {', '.join(missing)}")

    df = df.copy()
    for col in ("budget", "actual"):
        df[col] = pd.to_numeric(df[col], errors="coerce")
        if df[col].isna().any():
            bad = df.index[df[col].isna()].tolist()
            raise ValidationError(f"Non-numeric value in '{col}' at row(s): {bad}")
        if (df[col] < 0).any():
            bad = df.index[df[col] < 0].tolist()
            raise ValidationError(f"Negative value in '{col}' at row(s): {bad}")

    dupes = df.duplicated(subset=["fund", "account"], keep=False)
    if dupes.any():
        rows = df.loc[dupes, ["fund", "account"]].to_dict("records")
        raise ValidationError(f"Duplicate fund/account line(s): {rows}")

    return df


# --------------------------------------------------------------------------- #
# Load into SQLite (a real relational ledger)
# --------------------------------------------------------------------------- #
def load_to_sqlite(df: pd.DataFrame, db_path: str = ":memory:") -> sqlite3.Connection:
    """Load the validated ledger into SQLite and return the open connection."""
    conn = sqlite3.connect(db_path)
    df.to_sql("ledger", conn, if_exists="replace", index=False)
    return conn


# --------------------------------------------------------------------------- #
# Analysis
# --------------------------------------------------------------------------- #
def compute_variance(conn: sqlite3.Connection) -> pd.DataFrame:
    """Compute per-line variance (dollar and percent) via SQL."""
    query = """
        SELECT
            fund,
            account,
            description,
            budget,
            actual,
            (actual - budget)                              AS variance,
            CASE WHEN budget = 0 THEN NULL
                 ELSE ROUND((actual - budget) * 100.0 / budget, 1)
            END                                            AS variance_pct,
            CASE WHEN budget = 0 THEN NULL
                 ELSE ROUND(actual * 100.0 / budget, 1)
            END                                            AS pct_used
        FROM ledger
        ORDER BY fund, account
    """
    return pd.read_sql_query(query, conn)


def flag_exceptions(df: pd.DataFrame, threshold: float = DEFAULT_WATCH_THRESHOLD) -> pd.DataFrame:
    """Add a 'status' column: OVER BUDGET, WATCH, or OK."""
    df = df.copy()

    def status(row: pd.Series) -> str:
        if row["actual"] > row["budget"]:
            return "OVER BUDGET"
        if row["budget"] > 0 and row["actual"] >= threshold * row["budget"]:
            return "WATCH"
        return "OK"

    df["status"] = df.apply(status, axis=1)
    return df


# Report order, worst first, so a reviewer reads exceptions before OK lines.
STATUS_ORDER = ["OVER BUDGET", "WATCH", "OK"]


def status_counts(detail: pd.DataFrame) -> dict[str, int]:
    """Count flagged lines per status, always including all known statuses."""
    counts = detail["status"].value_counts().to_dict()
    return {status: int(counts.get(status, 0)) for status in STATUS_ORDER}


def fund_rollup(conn: sqlite3.Connection) -> pd.DataFrame:
    """Summarize budget, actual, and variance grouped by fund."""
    query = """
        SELECT
            fund,
            SUM(budget)            AS budget,
            SUM(actual)            AS actual,
            SUM(actual - budget)   AS variance
        FROM ledger
        GROUP BY fund
        ORDER BY fund
    """
    return pd.read_sql_query(query, conn)


@dataclass
class Summary:
    line_count: int
    total_budget: float
    total_actual: float

    @property
    def net_variance(self) -> float:
        return self.total_actual - self.total_budget

    @property
    def net_variance_pct(self) -> float:
        if self.total_budget == 0:
            return 0.0
        return self.net_variance * 100.0 / self.total_budget


def summarize(df: pd.DataFrame) -> Summary:
    return Summary(
        line_count=len(df),
        total_budget=float(df["budget"].sum()),
        total_actual=float(df["actual"].sum()),
    )


# --------------------------------------------------------------------------- #
# Reporting
# --------------------------------------------------------------------------- #
def _money(value: float) -> str:
    """Format a signed dollar amount as -$1,234.56 / $1,234.56."""
    return f"-${abs(value):,.2f}" if value < 0 else f"${value:,.2f}"


def format_fund_rollup(rollup: pd.DataFrame) -> str:
    """Render the per-fund budget/actual/variance rollup as an aligned text table."""
    lines = [
        "FUND ROLLUP",
        f"  {'Fund':<26}{'Budget':>16}{'Actual':>16}{'Variance':>16}",
    ]
    for _, r in rollup.iterrows():
        lines.append(
            f"  {str(r['fund'])[:26]:<26}"
            f"{_money(r['budget']):>16}"
            f"{_money(r['actual']):>16}"
            f"{_money(r['variance']):>16}"
        )
    return "\n".join(lines)


def print_report(detail: pd.DataFrame, rollup: pd.DataFrame, summary: Summary) -> None:
    direction = "under" if summary.net_variance <= 0 else "over"
    print("\n=== District Budget Variance Summary ===")
    print(
        f"Lines analyzed: {summary.line_count}   |   "
        f"Total budget: ${summary.total_budget:,.2f}   |   "
        f"Total actual: ${summary.total_actual:,.2f}"
    )
    print(
        f"Net variance: ${abs(summary.net_variance):,.2f} {direction} budget "
        f"({abs(summary.net_variance_pct):.2f}%)"
    )

    counts = status_counts(detail)
    print(
        "Status breakdown: "
        f"{counts['OVER BUDGET']} OVER BUDGET | "
        f"{counts['WATCH']} WATCH | "
        f"{counts['OK']} OK\n"
    )

    print(format_fund_rollup(rollup))
    print()

    flagged = detail[detail["status"] != "OK"]
    if flagged.empty:
        print("No exceptions: every account is within budget parameters.")
        return

    print("FLAGGED FOR REVIEW")
    for _, r in flagged.iterrows():
        line = f"{r['fund']} / {r['account']} {r['description']}"
        if r["status"] == "OVER BUDGET":
            print(f"  [OVER BUDGET] {line}  +${r['variance']:,.2f} ({r['variance_pct']:+.1f}%)")
        else:
            print(f"  [WATCH]       {line}  {r['pct_used']:.1f}% of budget used")
    print()


def write_reports(detail: pd.DataFrame, rollup: pd.DataFrame, out_dir: str) -> Path:
    """Write CSV + formatted Excel report. Returns the Excel path."""
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)

    csv_path = out / "budget_report.csv"
    detail.to_csv(csv_path, index=False)

    xlsx_path = out / "budget_report.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        detail.to_excel(writer, sheet_name="Line Detail", index=False)
        rollup.to_excel(writer, sheet_name="Fund Rollup", index=False)
        _highlight_flags(writer, detail)

    return xlsx_path


def _highlight_flags(writer: "pd.ExcelWriter", detail: pd.DataFrame) -> None:
    """Conditional highlighting: red for OVER BUDGET, amber for WATCH."""
    from openpyxl.styles import PatternFill

    ws = writer.sheets["Line Detail"]
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    amber = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    status_col = list(detail.columns).index("status") + 1  # 1-based
    for offset, status in enumerate(detail["status"], start=2):  # row 1 = header
        if status == "OVER BUDGET":
            ws.cell(row=offset, column=status_col).fill = red
        elif status == "WATCH":
            ws.cell(row=offset, column=status_col).fill = amber

    # Auto-size columns for readability.
    for col_cells in ws.columns:
        width = max(len(str(c.value)) for c in col_cells if c.value is not None)
        ws.column_dimensions[col_cells[0].column_letter].width = width + 2


# --------------------------------------------------------------------------- #
# Orchestration / CLI
# --------------------------------------------------------------------------- #
def analyze(input_csv: str, out_dir: str, threshold: float) -> tuple[pd.DataFrame, Summary]:
    raw = pd.read_csv(input_csv)
    clean = validate(raw)
    conn = load_to_sqlite(clean)
    try:
        detail = flag_exceptions(compute_variance(conn), threshold=threshold)
        rollup = fund_rollup(conn)
    finally:
        conn.close()

    summary = summarize(clean)
    print_report(detail, rollup, summary)
    path = write_reports(detail, rollup, out_dir)
    print(f"Report written to {path}")
    return detail, summary


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Analyze district budget vs. actuals and flag exceptions."
    )
    parser.add_argument("--input", "-i", required=True, help="Path to budget CSV.")
    parser.add_argument("--out", "-o", default="reports", help="Output directory.")
    parser.add_argument(
        "--threshold",
        "-t",
        type=float,
        default=DEFAULT_WATCH_THRESHOLD,
        help="WATCH threshold as a fraction of budget (default 0.90).",
    )
    parser.add_argument(
        "--fail-on-over",
        action="store_true",
        help="Exit non-zero (2) if any line is OVER BUDGET, for CI/automation gating.",
    )
    args = parser.parse_args(argv)

    try:
        detail, _ = analyze(args.input, args.out, args.threshold)
    except (ValidationError, FileNotFoundError) as exc:
        print(f"ERROR: {exc}")
        return 1

    if args.fail_on_over and status_counts(detail)["OVER BUDGET"] > 0:
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
